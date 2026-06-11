import hashlib
import hmac
import io
import json
from decimal import Decimal, InvalidOperation

from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.db import transaction
from django.db.models import Count, DecimalField, ExpressionWrapper, F, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_exempt

from django.conf import settings

from .auth_helpers import (
    create_token,
    get_token_user,
    login_user,
    register_user,
    require_user,
    user_payload,
)
from .models import Category, CustomerOrder, OrderItem, OrderStatusEvent, Payment, Product

MIN_QUANTITY_KG = Decimal("0.25")
MAX_QUANTITY_KG = Decimal("100")


# ---------------------------------------------------------------------------
# Helpers — HTTP responses
# ---------------------------------------------------------------------------

def json_response(data, status=200):
    return JsonResponse(data, status=status)


def binary_file_response(content, content_type, filename):
    response = HttpResponse(content, content_type=content_type)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# Helpers — Serialisers
# ---------------------------------------------------------------------------

def product_payload(product):
    return {
        "id": product.id,
        "name": product.name,
        "tamilName": product.tamil_name,
        "description": product.description,
        "tamilDescription": product.tamil_description,
        "price": float(product.price_per_kg),
        "image": product.image_url,
    }


def category_payload(category):
    return {
        "id": str(category.id),
        "icon": category.icon,
        "name": category.name,
        "tamilName": category.tamil_name,
        "description": category.description,
        "tamilDescription": category.tamil_description,
        "products": [product_payload(p) for p in category.products.filter(is_available=True)],
    }


def payment_payload(payment):
    return {
        "id": payment.id,
        "method": payment.method,
        "status": payment.status,
        "amount": float(payment.amount),
        "razorpay_order_id": payment.razorpay_order_id,
        "razorpay_payment_id": payment.razorpay_payment_id,
        "created_at": payment.created_at.isoformat(),
    }


def order_payload(order, include_sensitive=False):
    payload = {
        "order_number": order.order_number,
        "customer_name": order.customer_name,
        "status": order.status,
        "order_total": float(order.order_total),
        "payment_method": order.payment_method,
        "payment_status": order.payment_status,
        "created_at": order.created_at.isoformat(),
        "updated_at": order.updated_at.isoformat(),
        "user_id": order.user_id,
        "items": [
            {
                "product_name": item.product_name,
                "tamil_name": item.tamil_name,
                "quantity_kg": float(item.quantity_kg),
                "unit_price": float(item.unit_price),
                "line_total": float(item.line_total),
            }
            for item in order.items.all()
        ],
        "payments": [payment_payload(p) for p in order.payments.all()],
        "events": [
            {
                "status": event.status,
                "message": event.message,
                "created_at": event.created_at.isoformat(),
            }
            for event in order.events.all()
        ],
    }
    if include_sensitive:
        payload.update(
            {
                "customer_phone": order.customer_phone,
                "customer_address": order.customer_address,
                "customer_notes": order.customer_notes,
                "tracking_pin": order.tracking_pin,
            }
        )
    return payload


# ---------------------------------------------------------------------------
# Helpers — Misc
# ---------------------------------------------------------------------------

def parse_body(request):
    try:
        return json.loads(request.body.decode("utf-8") or "{}")
    except (json.JSONDecodeError, UnicodeDecodeError):
        return {}


def require_admin(request):
    user = get_token_user(request)
    return bool(user and user.is_staff)


def sanitize_text(value, max_length=500):
    return str(value or "").strip()[:max_length]


# ---------------------------------------------------------------------------
# Helpers — Order validation
# ---------------------------------------------------------------------------

def validate_order_items(items):
    """Validate and price order items against the database.

    Returns (validated_list, total, error_string).
    On success error_string is None; on failure the first two values are None.
    """
    if not items:
        return [], Decimal("0"), None

    validated = []
    total = Decimal("0")

    for item in items:
        product_id = item.get("product_id")
        product = Product.objects.filter(id=product_id, is_available=True).first()
        if not product:
            return None, None, f"Product {product_id} is not available."

        try:
            quantity = Decimal(str(item.get("quantity_kg", "1")))
        except (InvalidOperation, TypeError):
            return None, None, "Invalid quantity."

        if quantity < MIN_QUANTITY_KG or quantity > MAX_QUANTITY_KG:
            return None, None, (
                f"Quantity must be between {MIN_QUANTITY_KG} and {MAX_QUANTITY_KG} kg."
            )

        unit_price = product.price_per_kg
        if unit_price <= 0:
            return None, None, (
                f"{product.name} requires price confirmation. Add it to notes instead."
            )

        total += quantity * unit_price
        validated.append(
            {
                "product": product,
                "product_name": product.name,
                "tamil_name": product.tamil_name,
                "quantity_kg": quantity,
                "unit_price": unit_price,
            }
        )

    return validated, total, None


# ---------------------------------------------------------------------------
# Helpers — Email
# ---------------------------------------------------------------------------

def _items_text(order):
    return "\n".join(
        f"  - {item.product_name} x {item.quantity_kg} kg = Rs.{item.line_total:.2f}"
        for item in order.items.all()
    )


def _send(subject, body, to):
    try:
        send_mail(subject, body, settings.DEFAULT_FROM_EMAIL, [to], fail_silently=True)
    except Exception:
        pass


STORE_FOOTER = (
    "\nFor queries: 70104 82463 | 83448 80228"
    "\nAnnai Health Foods, Madakkudi, Pallividai, Samayapuram, Trichy-621 112"
    "\nFSSAI Lic. No.: 22420308000104"
)


def send_order_receipt(order):
    """Receipt email to customer + new-order alert to admin."""
    if order.customer_email:
        _send(
            subject=f"Order Received \u2013 {order.order_number} | ANNAI HEALTH MASALA",
            body=(
                f"Dear {order.customer_name},\n\n"
                f"Thank you for your order!\n\n"
                f"Order Number : {order.order_number}\n"
                f"Tracking PIN : {order.tracking_pin}\n"
                f"Date         : {order.created_at.strftime('%d %b %Y, %I:%M %p')}\n\n"
                f"Items Ordered:\n{_items_text(order)}\n\n"
                f"Order Total  : Rs.{order.order_total:.2f}\n"
                f"Payment      : {order.get_payment_method_display()}\n"
                f"Status       : {order.get_payment_status_display()}\n\n"
                f"Delivery Address:\n{order.customer_address}\n\n"
                f"Track your order using Order Number + Phone or Tracking PIN."
                f"{STORE_FOOTER}\n\nThank you for choosing ANNAI HEALTH MASALA!"
            ),
            to=order.customer_email,
        )

    admin_email = getattr(settings, "ADMIN_EMAIL", "")
    if admin_email:
        _send(
            subject=f"New Order {order.order_number} \u2013 Rs.{order.order_total:.2f}",
            body=(
                f"New order received.\n\n"
                f"Order  : {order.order_number}\n"
                f"Name   : {order.customer_name}\n"
                f"Phone  : {order.customer_phone}\n"
                f"Email  : {order.customer_email}\n"
                f"Total  : Rs.{order.order_total:.2f}\n"
                f"Method : {order.get_payment_method_display()}\n\n"
                f"Items:\n{_items_text(order)}\n\n"
                f"Address: {order.customer_address}"
            ),
            to=admin_email,
        )


def send_payment_confirmation(order):
    """Full paid receipt to customer."""
    if not order.customer_email:
        return
    _send(
        subject=f"Payment Confirmed \u2013 {order.order_number} | ANNAI HEALTH MASALA",
        body=(
            f"Dear {order.customer_name},\n\n"
            f"Your payment has been received and confirmed.\n\n"
            f"Order Number  : {order.order_number}\n"
            f"Tracking PIN  : {order.tracking_pin}\n"
            f"Date          : {order.created_at.strftime('%d %b %Y, %I:%M %p')}\n\n"
            f"Items Ordered:\n{_items_text(order)}\n\n"
            f"Order Total   : Rs.{order.order_total:.2f}\n"
            f"Payment Mode  : {order.get_payment_method_display()}\n"
            f"Payment Status: PAID\n\n"
            f"Delivery Address:\n{order.customer_address}\n\n"
            f"Your order is being processed. We will notify you at every step."
            f"{STORE_FOOTER}\n\nThank you for choosing ANNAI HEALTH MASALA!"
        ),
        to=order.customer_email,
    )


def send_status_update_email(order, message):
    """Status-change / admin-message notification to customer."""
    if not order.customer_email:
        return
    _send(
        subject=f"Order Update \u2013 {order.order_number} | ANNAI HEALTH MASALA",
        body=(
            f"Dear {order.customer_name},\n\n"
            f"Order Number   : {order.order_number}\n"
            f"Current Status : {order.get_status_display()}\n"
            f"Payment Status : {order.get_payment_status_display()}\n\n"
            f"Message from us:\n{message}\n\n"
            f"Track your order using Order Number + Phone or Tracking PIN: {order.tracking_pin}"
            f"{STORE_FOOTER}"
        ),
        to=order.customer_email,
    )


# ---------------------------------------------------------------------------
# Helpers — Razorpay
# ---------------------------------------------------------------------------

def razorpay_client():
    if not settings.RAZORPAY_KEY_ID or not settings.RAZORPAY_KEY_SECRET:
        return None
    try:
        import razorpay
        return razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))
    except ImportError:
        return None


def create_razorpay_order(order, payment_method=CustomerOrder.PAYMENT_ONLINE):
    client = razorpay_client()
    if not client:
        return None, "Online payment is not configured."

    amount_paise = int(order.order_total * 100)
    if amount_paise < 100:
        return None, "Order total is too low for online payment."

    stored_method = (
        payment_method
        if payment_method in CustomerOrder.ONLINE_PAYMENT_METHODS
        else CustomerOrder.PAYMENT_ONLINE
    )

    try:
        rz_order = client.order.create(
            {
                "amount": amount_paise,
                "currency": "INR",
                "receipt": order.order_number,
                "notes": {"order_number": order.order_number, "payment_method": stored_method},
            }
        )
        payment = Payment.objects.create(
            order=order,
            method=stored_method,
            status=CustomerOrder.PAYMENT_PENDING,
            amount=order.order_total,
            razorpay_order_id=rz_order["id"],
        )
        return {
            "razorpay_order_id": rz_order["id"],
            "razorpay_key_id": settings.RAZORPAY_KEY_ID,
            "amount": amount_paise,
            "currency": "INR",
            "payment_id": payment.id,
            "payment_method": stored_method,
        }, None
    except Exception as exc:
        return None, f"Payment gateway error: {exc}"


# ---------------------------------------------------------------------------
# Public API views
# ---------------------------------------------------------------------------

def menu(request):
    if request.method != "GET":
        return json_response({"error": "GET required"}, status=405)
    categories = Category.objects.prefetch_related("products").all()
    return json_response({"categories": [category_payload(c) for c in categories]})


def payment_config(request):
    if request.method != "GET":
        return json_response({"error": "GET required"}, status=405)
    client = razorpay_client()
    return json_response(
        {
            "online_enabled": bool(client),
            "razorpay_key_id": settings.RAZORPAY_KEY_ID if client else "",
            "currency": "INR",
        }
    )


@csrf_exempt
def register(request):
    if request.method != "POST":
        return json_response({"error": "POST required"}, status=405)
    user, error = register_user(parse_body(request))
    if error:
        return json_response({"error": error}, status=400)
    token = create_token(user)
    return json_response({"user": user_payload(user), "token": token.key}, status=201)


@csrf_exempt
def login(request):
    if request.method != "POST":
        return json_response({"error": "POST required"}, status=405)
    user, error = login_user(parse_body(request))
    if error:
        return json_response({"error": error}, status=401)
    token = create_token(user)
    return json_response({"user": user_payload(user), "token": token.key})


@csrf_exempt
def logout(request):
    if request.method != "POST":
        return json_response({"error": "POST required"}, status=405)
    user = get_token_user(request)
    if user:
        user.auth_tokens.all().delete()
    return json_response({"ok": True})


def me(request):
    user, error_response = require_user(request)
    if error_response:
        return error_response
    return json_response({"user": user_payload(user)})


@csrf_exempt
def create_order(request):
    if request.method != "POST":
        return json_response({"error": "POST required"}, status=405)

    data = parse_body(request)
    customer = data.get("customer", {})
    items = data.get("items", [])
    payment_method = data.get("payment_method", CustomerOrder.PAYMENT_COD)

    valid_methods = {choice[0] for choice in CustomerOrder.PAYMENT_METHOD_CHOICES}
    if payment_method not in valid_methods:
        return json_response({"error": "Invalid payment method."}, status=400)

    required_fields = ["name", "phone", "address"]
    if any(not sanitize_text(customer.get(field)) for field in required_fields):
        return json_response(
            {"error": "Customer name, phone, and address are required."}, status=400
        )
    if not items and not sanitize_text(customer.get("notes")):
        return json_response({"error": "Add an item or request note."}, status=400)

    validated_items, order_total, item_error = validate_order_items(items)
    if item_error:
        return json_response({"error": item_error}, status=400)

    is_online = payment_method in CustomerOrder.ONLINE_PAYMENT_METHODS
    if is_online and not razorpay_client():
        return json_response(
            {"error": "Online payment is not available. Choose Cash on Delivery."}, status=400
        )

    auth_user, error_response = require_user(request)
    if error_response:
        return error_response

    with transaction.atomic():
        order = CustomerOrder.objects.create(
            user=auth_user,
            customer_name=sanitize_text(customer["name"], 120),
            customer_phone=sanitize_text(customer["phone"], 30),
            customer_email=sanitize_text(customer.get("email", ""), 254),
            customer_address=sanitize_text(customer["address"], 1000),
            customer_notes=sanitize_text(customer.get("notes"), 2000),
            language=data.get("language", "en"),
            order_total=order_total,
            payment_method=payment_method,
            payment_status=CustomerOrder.PAYMENT_PENDING,
        )
        for item in validated_items:
            OrderItem.objects.create(
                order=order,
                product=item["product"],
                product_name=item["product_name"],
                tamil_name=item["tamil_name"],
                quantity_kg=item["quantity_kg"],
                unit_price=item["unit_price"],
            )
        if not is_online:
            Payment.objects.create(
                order=order,
                method=CustomerOrder.PAYMENT_COD,
                status=CustomerOrder.PAYMENT_PENDING,
                amount=order_total,
            )
        OrderStatusEvent.objects.create(
            order=order,
            status=CustomerOrder.STATUS_RECEIVED,
            message="Your order has been received. We will confirm it shortly.",
        )

    response_data = {"order": order_payload(order, include_sensitive=True)}

    if is_online:
        razorpay_data, payment_error = create_razorpay_order(order, payment_method)
        if payment_error:
            order.delete()
            return json_response({"error": payment_error}, status=400)
        response_data["payment"] = razorpay_data

    send_order_receipt(order)
    return json_response(response_data, status=201)


def order_detail(request, order_number):
    if request.method != "GET":
        return json_response({"error": "GET required"}, status=405)

    phone = sanitize_text(request.GET.get("phone"))
    pin = sanitize_text(request.GET.get("pin"))
    order = get_object_or_404(
        CustomerOrder.objects.prefetch_related("items", "events", "payments"),
        order_number=order_number,
    )

    auth_user = get_token_user(request)
    is_owner = bool(auth_user and order.user_id == auth_user.id)
    is_admin = require_admin(request)
    phone_match = bool(phone and phone.replace(" ", "") == order.customer_phone.replace(" ", ""))
    pin_match = bool(pin and pin == order.tracking_pin)

    if not (is_owner or is_admin or phone_match or pin_match):
        return json_response(
            {"error": "Provide the phone number or tracking PIN used when placing the order."},
            status=403,
        )

    return json_response(
        {"order": order_payload(order, include_sensitive=is_owner or is_admin or pin_match)}
    )


def my_orders(request):
    if request.method != "GET":
        return json_response({"error": "GET required"}, status=405)
    user, error_response = require_user(request)
    if error_response:
        return error_response
    orders = (
        CustomerOrder.objects.filter(user=user)
        .prefetch_related("items", "events", "payments")
        .order_by("-created_at")[:50]
    )
    return json_response(
        {"orders": [order_payload(order, include_sensitive=True) for order in orders]}
    )


# ---------------------------------------------------------------------------
# Payment views
# ---------------------------------------------------------------------------

@csrf_exempt
def razorpay_order_view(request):
    """POST /api/payments/create-order/ — create Razorpay order from server-side amount."""
    if request.method != "POST":
        return json_response({"error": "POST required"}, status=405)

    data = parse_body(request)
    annai_order_number = sanitize_text(data.get("annai_order_number"))
    if not annai_order_number:
        return json_response({"error": "annai_order_number required"}, status=400)

    order = CustomerOrder.objects.filter(order_number=annai_order_number).first()
    if not order:
        return json_response({"error": "Order not found"}, status=404)

    amount_paise = int(order.order_total * 100)
    if amount_paise <= 0 or amount_paise > 100_000_000:
        return json_response({"error": "Invalid order amount"}, status=400)

    client = razorpay_client()
    if not client:
        return json_response({"error": "Online payment not configured"}, status=503)

    try:
        rz_order = client.order.create(
            {
                "amount": amount_paise,
                "currency": "INR",
                "receipt": annai_order_number,
                "notes": {"shop": "Annai Health Masala"},
                "payment_capture": 1,
            }
        )
    except Exception as exc:
        return json_response({"error": f"Gateway error: {exc}"}, status=502)

    order.razorpay_order_id = rz_order["id"]
    order.save(update_fields=["razorpay_order_id", "updated_at"])
    Payment.objects.filter(order=order).update(razorpay_order_id=rz_order["id"])

    return json_response({"razorpay_order_id": rz_order["id"], "amount": amount_paise})


@csrf_exempt
def verify_payment(request):
    if request.method != "POST":
        return json_response({"error": "POST required"}, status=405)

    data = parse_body(request)
    order_number = sanitize_text(data.get("order_number") or data.get("annai_order_number"))
    razorpay_order_id = sanitize_text(data.get("razorpay_order_id"))
    razorpay_payment_id = sanitize_text(data.get("razorpay_payment_id"))
    razorpay_signature = sanitize_text(data.get("razorpay_signature"), 128)

    if not all([order_number, razorpay_order_id, razorpay_payment_id, razorpay_signature]):
        return json_response({"error": "Missing payment verification fields."}, status=400)

    order = get_object_or_404(CustomerOrder, order_number=order_number)
    payment = Payment.objects.filter(order=order, razorpay_order_id=razorpay_order_id).first()
    if not payment:
        return json_response({"error": "Payment record not found."}, status=404)

    if payment.status == CustomerOrder.PAYMENT_PAID:
        order = CustomerOrder.objects.prefetch_related("items", "events").get(pk=order.pk)
        return json_response({"order": order_payload(order, include_sensitive=True)})

    if not settings.RAZORPAY_KEY_SECRET:
        return json_response({"error": "Payment verification is not configured."}, status=400)

    expected = hmac.new(
        settings.RAZORPAY_KEY_SECRET.encode(),
        f"{razorpay_order_id}|{razorpay_payment_id}".encode(),
        hashlib.sha256,
    ).hexdigest()

    if not hmac.compare_digest(expected, razorpay_signature):
        payment.status = CustomerOrder.PAYMENT_FAILED
        payment.save(update_fields=["status"])
        order.payment_status = CustomerOrder.PAYMENT_FAILED
        order.save(update_fields=["payment_status", "updated_at"])
        return json_response({"error": "Payment verification failed."}, status=400)

    payment.status = CustomerOrder.PAYMENT_PAID
    payment.razorpay_payment_id = razorpay_payment_id
    payment.razorpay_signature = razorpay_signature
    payment.save(update_fields=["status", "razorpay_payment_id", "razorpay_signature"])

    order.payment_status = CustomerOrder.PAYMENT_PAID
    order.razorpay_payment_id = razorpay_payment_id
    order.save(update_fields=["payment_status", "razorpay_payment_id", "updated_at"])

    OrderStatusEvent.objects.create(
        order=order,
        status=order.status,
        message="Payment received successfully. Thank you!",
    )
    order = CustomerOrder.objects.prefetch_related("items", "events").get(pk=order.pk)
    return json_response({"order": order_payload(order, include_sensitive=True)})


@csrf_exempt
def webhook(request):
    """POST /api/payments/webhook/ — Razorpay event webhook."""
    if request.method != "POST":
        return json_response({"error": "POST required"}, status=405)

    secret = settings.RAZORPAY_KEY_SECRET
    if secret:
        sig = request.headers.get("X-Razorpay-Signature", "")
        expected = hmac.new(secret.encode(), request.body, hashlib.sha256).hexdigest()
        if not hmac.compare_digest(expected, sig):
            return json_response({"error": "Invalid signature"}, status=400)

    try:
        event = json.loads(request.body)
    except (json.JSONDecodeError, UnicodeDecodeError):
        return json_response({"status": "ok"})

    if event.get("event") == "payment.captured":
        entity = event.get("payload", {}).get("payment", {}).get("entity", {})
        rz_order_id = entity.get("order_id", "")
        rz_payment_id = entity.get("id", "")
        if rz_order_id:
            CustomerOrder.objects.filter(razorpay_order_id=rz_order_id).update(
                payment_status=CustomerOrder.PAYMENT_PAID,
                razorpay_payment_id=rz_payment_id,
            )
            Payment.objects.filter(razorpay_order_id=rz_order_id).update(
                status=CustomerOrder.PAYMENT_PAID,
                razorpay_payment_id=rz_payment_id,
            )

    return json_response({"status": "ok"})


# ---------------------------------------------------------------------------
# Admin views
# ---------------------------------------------------------------------------

def admin_dashboard(request):
    if request.method != "GET":
        return json_response({"error": "GET required"}, status=405)
    if not require_admin(request):
        return json_response({"error": "Invalid admin credentials."}, status=403)

    total_orders = CustomerOrder.objects.count()
    active_orders = CustomerOrder.objects.exclude(
        status__in=[CustomerOrder.STATUS_DELIVERED, CustomerOrder.STATUS_CANCELLED]
    ).count()
    total_users = User.objects.filter(is_staff=False).count()
    total_revenue = (
        Payment.objects.filter(status=CustomerOrder.PAYMENT_PAID)
        .aggregate(total=Sum("amount"))["total"]
        or 0
    )
    pending_payments = CustomerOrder.objects.filter(
        payment_status=CustomerOrder.PAYMENT_PENDING
    ).count()

    orders_by_status = {
        status: CustomerOrder.objects.filter(status=status).count()
        for status, _ in CustomerOrder.STATUS_CHOICES
    }

    line_total_expr = ExpressionWrapper(
        F("quantity_kg") * F("unit_price"),
        output_field=DecimalField(max_digits=12, decimal_places=2),
    )
    product_sales = (
        OrderItem.objects.values("product_name")
        .annotate(
            total_kg=Sum("quantity_kg"),
            revenue=Sum(line_total_expr),
            order_count=Count("order_id", distinct=True),
        )
        .order_by("-revenue")[:15]
    )

    recent_orders = (
        CustomerOrder.objects.prefetch_related("items").order_by("-created_at")[:12]
    )

    return json_response(
        {
            "stats": {
                "total_orders": total_orders,
                "active_orders": active_orders,
                "total_users": total_users,
                "total_revenue": float(total_revenue),
                "pending_payments": pending_payments,
            },
            "orders_by_status": orders_by_status,
            "product_sales": [
                {
                    "product_name": row["product_name"],
                    "total_kg": float(row["total_kg"] or 0),
                    "revenue": float(row["revenue"] or 0),
                    "order_count": row["order_count"],
                }
                for row in product_sales
            ],
            "recent_orders": [
                order_payload(order, include_sensitive=True) for order in recent_orders
            ],
        }
    )


def admin_orders(request):
    if request.method != "GET":
        return json_response({"error": "GET required"}, status=405)
    if not require_admin(request):
        return json_response({"error": "Invalid admin credentials."}, status=403)
    orders = (
        CustomerOrder.objects.prefetch_related("items", "events", "payments")
        .order_by("-created_at")[:100]
    )
    return json_response(
        {"orders": [order_payload(order, include_sensitive=True) for order in orders]}
    )


def admin_payments(request):
    if request.method != "GET":
        return json_response({"error": "GET required"}, status=405)
    if not require_admin(request):
        return json_response({"error": "Invalid admin credentials."}, status=403)

    payments = Payment.objects.select_related("order").order_by("-created_at")[:100]
    return json_response(
        {
            "payments": [
                {
                    **payment_payload(p),
                    "order_number": p.order.order_number,
                    "customer_name": p.order.customer_name,
                    "customer_phone": p.order.customer_phone,
                }
                for p in payments
            ]
        }
    )


def admin_users(request):
    if request.method != "GET":
        return json_response({"error": "GET required"}, status=405)
    if not require_admin(request):
        return json_response({"error": "Invalid admin credentials."}, status=403)

    users = (
        User.objects.filter(is_staff=False)
        .annotate(order_count=Count("orders"))
        .select_related("profile")
        .order_by("-date_joined")[:200]
    )
    return json_response(
        {
            "users": [
                {**user_payload(u), "order_count": u.order_count}
                for u in users
            ]
        }
    )


def admin_user_detail(request, user_id):
    if request.method != "GET":
        return json_response({"error": "GET required"}, status=405)
    if not require_admin(request):
        return json_response({"error": "Invalid admin credentials."}, status=403)

    user = get_object_or_404(User.objects.select_related("profile"), pk=user_id)
    orders = (
        CustomerOrder.objects.filter(user=user)
        .prefetch_related("items", "events", "payments")
        .order_by("-created_at")
    )
    payments = (
        Payment.objects.filter(order__user=user)
        .select_related("order")
        .order_by("-created_at")
    )
    return json_response(
        {
            "user": {**user_payload(user), "order_count": orders.count()},
            "orders": [order_payload(order, include_sensitive=True) for order in orders],
            "payments": [
                {**payment_payload(p), "order_number": p.order.order_number}
                for p in payments
            ],
        }
    )


@csrf_exempt
def admin_notify(request, order_number):
    if request.method != "POST":
        return json_response({"error": "POST required"}, status=405)
    if not require_admin(request):
        return json_response({"error": "Invalid admin credentials."}, status=403)

    data = parse_body(request)
    message = sanitize_text(data.get("message"), 500)
    payment_status = data.get("payment_status")
    valid_payment_statuses = {choice[0] for choice in CustomerOrder.PAYMENT_STATUS_CHOICES}

    if not message and not payment_status:
        return json_response({"error": "Message or payment_status is required."}, status=400)

    order = get_object_or_404(CustomerOrder, order_number=order_number)

    if payment_status:
        if payment_status not in valid_payment_statuses:
            return json_response({"error": "Invalid payment status."}, status=400)
        order.payment_status = payment_status
        order.save(update_fields=["payment_status", "updated_at"])
        latest_payment = order.payments.order_by("-created_at").first()
        if latest_payment:
            latest_payment.status = payment_status
            latest_payment.save(update_fields=["status"])
        if not message:
            message = "Your payment has been received. Thank you!"
        if payment_status == CustomerOrder.PAYMENT_PAID:
            order.refresh_from_db()
            send_payment_confirmation(order)

    OrderStatusEvent.objects.create(order=order, status=order.status, message=message)
    send_status_update_email(order, message)

    order = get_object_or_404(
        CustomerOrder.objects.prefetch_related("items", "events", "payments"),
        order_number=order_number,
    )
    return json_response({"order": order_payload(order, include_sensitive=True)})


@csrf_exempt
def update_order_status(request, order_number):
    if request.method != "PATCH":
        return json_response({"error": "PATCH required"}, status=405)
    if not require_admin(request):
        return json_response({"error": "Invalid admin credentials."}, status=403)

    data = parse_body(request)
    status = data.get("status")
    valid_statuses = {choice[0] for choice in CustomerOrder.STATUS_CHOICES}
    if status not in valid_statuses:
        return json_response({"error": "Invalid status."}, status=400)

    order = get_object_or_404(CustomerOrder, order_number=order_number)
    order.status = status
    order.save(update_fields=["status", "updated_at"])

    notify_message = (
        sanitize_text(data.get("message"), 500)
        or f"Your order status is now {order.get_status_display()}."
    )
    OrderStatusEvent.objects.create(order=order, status=status, message=notify_message)
    send_status_update_email(order, notify_message)

    order = get_object_or_404(
        CustomerOrder.objects.prefetch_related("items", "events"),
        order_number=order_number,
    )
    return json_response({"order": order_payload(order, include_sensitive=True)})


# ---------------------------------------------------------------------------
# Export views
# ---------------------------------------------------------------------------

def admin_export_excel(request):
    """GET /api/admin/export/excel/ — download full data as .xlsx"""
    if request.method != "GET":
        return json_response({"error": "GET required"}, status=405)
    if not require_admin(request):
        return json_response({"error": "Invalid admin credentials."}, status=403)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return json_response(
            {"error": "openpyxl not installed. Run: pip install -r requirements.txt"},
            status=500,
        )

    try:
        wb = openpyxl.Workbook()

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="12633D")
        center = Alignment(horizontal="center")

        def style_header(ws, headers):
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center

        ws_orders = wb.active
        ws_orders.title = "Orders"
        style_header(ws_orders, [
            "Order Number", "Date", "Customer Name", "Phone", "Email",
            "Address", "Items", "Order Total (Rs)", "Payment Method",
            "Payment Status", "Order Status", "Tracking PIN", "Notes",
        ])
        orders = CustomerOrder.objects.prefetch_related("items").order_by("-created_at")
        for o in orders:
            items_str = "; ".join(
                f"{i.product_name} x{i.quantity_kg}kg" for i in o.items.all()
            )
            ws_orders.append([
                o.order_number,
                o.created_at.strftime("%d-%m-%Y %H:%M"),
                o.customer_name,
                o.customer_phone,
                o.customer_email,
                o.customer_address,
                items_str,
                float(o.order_total),
                o.get_payment_method_display(),
                o.get_payment_status_display(),
                o.get_status_display(),
                o.tracking_pin,
                o.customer_notes,
            ])

        ws_pay = wb.create_sheet("Payments")
        style_header(ws_pay, [
            "Payment ID", "Order Number", "Customer Name", "Phone",
            "Method", "Status", "Amount (Rs)", "Razorpay Order ID",
            "Razorpay Payment ID", "Date",
        ])
        for p in Payment.objects.select_related("order").order_by("-created_at"):
            ws_pay.append([
                p.id,
                p.order.order_number,
                p.order.customer_name,
                p.order.customer_phone,
                p.method,
                p.status,
                float(p.amount),
                p.razorpay_order_id,
                p.razorpay_payment_id,
                p.created_at.strftime("%d-%m-%Y %H:%M"),
            ])

        ws_users = wb.create_sheet("Customers")
        style_header(ws_users, [
            "User ID", "Username", "Full Name", "Email", "Phone",
            "Address", "Total Orders", "Joined",
        ])
        for u in (
            User.objects.filter(is_staff=False)
            .annotate(order_count=Count("orders"))
            .select_related("profile")
            .order_by("-date_joined")
        ):
            profile = getattr(u, "profile", None)
            ws_users.append([
                u.id,
                u.username,
                u.get_full_name() or u.first_name,
                u.email,
                profile.phone if profile else "",
                profile.address if profile else "",
                u.order_count,
                u.date_joined.strftime("%d-%m-%Y"),
            ])

        ws_logs = wb.create_sheet("Order Logs")
        style_header(ws_logs, [
            "Order Number", "Customer", "Status", "Message", "Timestamp",
        ])
        for event in OrderStatusEvent.objects.select_related("order").order_by("-created_at"):
            ws_logs.append([
                event.order.order_number,
                event.order.customer_name,
                event.status,
                event.message,
                event.created_at.strftime("%d-%m-%Y %H:%M"),
            ])

        for ws in wb.worksheets:
            for col in ws.columns:
                if not col:
                    continue
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buf = io.BytesIO()
        wb.save(buf)
        return binary_file_response(
            buf.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "annai-health-report.xlsx",
        )
    except Exception as exc:
        return json_response({"error": f"Excel export failed: {exc}"}, status=500)


def admin_export_receipt_pdf(request, order_number):
    """GET /api/admin/export/receipt/<order_number>/ — download order receipt as PDF"""
    if request.method != "GET":
        return json_response({"error": "GET required"}, status=405)
    if not require_admin(request):
        return json_response({"error": "Invalid admin credentials."}, status=403)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    except ImportError:
        return json_response(
            {"error": "reportlab not installed. Run: pip install -r requirements.txt"},
            status=500,
        )

    order = (
        CustomerOrder.objects.prefetch_related("items", "events", "payments")
        .filter(order_number=order_number)
        .first()
    )
    if not order:
        return json_response({"error": "Order not found."}, status=404)

    try:
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            rightMargin=2 * cm, leftMargin=2 * cm,
            topMargin=2 * cm, bottomMargin=2 * cm,
        )

        styles = getSampleStyleSheet()
        green = colors.HexColor("#12633D")
        light = colors.HexColor("#f7f4ee")
        title_style = ParagraphStyle(
            "title", parent=styles["Title"], textColor=green, fontSize=20, spaceAfter=4,
        )
        sub_style = ParagraphStyle(
            "sub", parent=styles["Normal"], textColor=green, fontSize=10, spaceAfter=2,
        )
        heading_style = ParagraphStyle(
            "heading", parent=styles["Heading2"], textColor=green,
            fontSize=12, spaceBefore=14, spaceAfter=6,
        )
        normal = styles["Normal"]

        story = [
            Paragraph("ANNAI HEALTH MASALA", title_style),
            Paragraph("Madakkudi, Pallividai, Samayapuram, Trichy - 621 112", sub_style),
            Paragraph("Phone: 70104 82463 | 83448 80228", sub_style),
            Paragraph("FSSAI Lic. No.: 22420308000104", sub_style),
            Spacer(1, 0.4 * cm),
            Table(
                [[""]], colWidths=[doc.width],
                style=TableStyle([("LINEBELOW", (0, 0), (-1, -1), 1.5, green)]),
            ),
            Spacer(1, 0.3 * cm),
            Paragraph("ORDER RECEIPT", ParagraphStyle(
                "receipt", parent=styles["Title"], textColor=green,
                fontSize=15, spaceAfter=10, alignment=TA_CENTER,
            )),
        ]

        meta = [
            ["Order Number", order.order_number, "Date", order.created_at.strftime("%d %b %Y, %I:%M %p")],
            ["Tracking PIN", order.tracking_pin, "Order Status", order.get_status_display()],
            ["Payment Mode", order.get_payment_method_display(), "Payment Status", order.get_payment_status_display()],
        ]
        meta_table = Table(meta, colWidths=[3.5 * cm, 7.5 * cm, 3.5 * cm, 5.5 * cm])
        meta_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ("TEXTCOLOR", (0, 0), (0, -1), green),
            ("TEXTCOLOR", (2, 0), (2, -1), green),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [light, colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("PADDING", (0, 0), (-1, -1), 5),
        ]))
        story.extend([meta_table, Spacer(1, 0.4 * cm)])

        story.append(Paragraph("Customer Details", heading_style))
        cust_data = [
            ["Name", order.customer_name],
            ["Phone", order.customer_phone],
            ["Email", order.customer_email or "—"],
            ["Address", order.customer_address],
        ]
        if order.customer_notes:
            cust_data.append(["Notes", order.customer_notes])
        cust_table = Table(cust_data, colWidths=[3.5 * cm, doc.width - 3.5 * cm])
        cust_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("TEXTCOLOR", (0, 0), (0, -1), green),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [light, colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("PADDING", (0, 0), (-1, -1), 5),
        ]))
        story.extend([cust_table, Spacer(1, 0.4 * cm)])

        story.append(Paragraph("Items Ordered", heading_style))
        item_rows = [["#", "Product", "Qty (kg)", "Unit Price (Rs)", "Line Total (Rs)"]]
        for idx, item in enumerate(order.items.all(), 1):
            item_rows.append([
                str(idx),
                item.product_name,
                f"{float(item.quantity_kg):.2f}",
                f"{float(item.unit_price):.2f}",
                f"{float(item.line_total):.2f}",
            ])
        item_rows.append(["", "", "", "TOTAL", f"Rs. {float(order.order_total):.2f}"])
        item_table = Table(item_rows, colWidths=[1 * cm, 8.5 * cm, 2.5 * cm, 3.5 * cm, 4.5 * cm])
        item_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), green),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (3, -1), (-1, -1), "Helvetica-Bold"),
            ("TEXTCOLOR", (3, -1), (-1, -1), green),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, light]),
            ("GRID", (0, 0), (-1, -2), 0.3, colors.lightgrey),
            ("LINEABOVE", (0, -1), (-1, -1), 1, green),
            ("PADDING", (0, 0), (-1, -1), 5),
        ]))
        story.extend([item_table, Spacer(1, 0.4 * cm)])

        pay_records = list(order.payments.all())
        if pay_records:
            story.append(Paragraph("Payment Records", heading_style))
            pay_rows = [["Payment ID", "Method", "Status", "Amount (Rs)", "Razorpay ID"]]
            for payment in pay_records:
                pay_rows.append([
                    str(payment.id), payment.method, payment.status,
                    f"{float(payment.amount):.2f}",
                    payment.razorpay_payment_id or "—",
                ])
            pay_table = Table(
                pay_rows,
                colWidths=[2 * cm, 3 * cm, 2.5 * cm, 3.5 * cm, doc.width - 11 * cm],
            )
            pay_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), green),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, light]),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
                ("PADDING", (0, 0), (-1, -1), 5),
            ]))
            story.extend([pay_table, Spacer(1, 0.4 * cm)])

        events = list(order.events.all())
        if events:
            story.append(Paragraph("Order Timeline", heading_style))
            ev_rows = [["Time", "Status", "Message"]]
            for event in events:
                ev_rows.append([
                    event.created_at.strftime("%d %b %Y %H:%M"),
                    event.status.title(),
                    event.message,
                ])
            ev_table = Table(ev_rows, colWidths=[4 * cm, 3 * cm, doc.width - 7 * cm])
            ev_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), green),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, light]),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
                ("PADDING", (0, 0), (-1, -1), 5),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            story.append(ev_table)

        story.extend([
            Spacer(1, 0.6 * cm),
            Table(
                [[""]], colWidths=[doc.width],
                style=TableStyle([("LINEABOVE", (0, 0), (-1, -1), 1, green)]),
            ),
            Paragraph(
                "Thank you for choosing ANNAI HEALTH MASALA! Your health is our priority.",
                ParagraphStyle(
                    "footer", parent=normal, textColor=green,
                    fontSize=9, alignment=TA_CENTER, spaceBefore=6,
                ),
            ),
        ])

        doc.build(story)
        return binary_file_response(
            buf.getvalue(),
            "application/pdf",
            f"receipt-{order.order_number}.pdf",
        )
    except Exception as exc:
        return json_response({"error": f"PDF export failed: {exc}"}, status=500)
